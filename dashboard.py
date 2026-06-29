"""
dashboard.py  –  PT Pinus Merah Abadi | FAD
Principal Payment Planning Dashboard

Jalankan: streamlit run dashboard.py
URL      : http://localhost:8501
"""

import json
import sys
import logging
import io
from datetime import datetime, date, timedelta
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st

# ──────────────────────────────────────────────────────────────────────────────
#  PATH  –  edit di sini kalau lokasi berubah
# ──────────────────────────────────────────────────────────────────────────────

DASHBOARD_DATA   = Path(__file__).parent / "data" / "Dashboard_Data.xlsx"
LAST_UPDATE_FILE = Path(__file__).parent / "data" / "last_update.json"
LOGO_PATH        = Path(__file__).parent / "LOGO.jpg"   # lokal
LOGO_PATH_CLOUD  = Path("/mount/src/pma-dashboard/LOGO.jpg")  # Streamlit Cloud
LOG_DIR          = Path(__file__).parent / "logs"

# ── Supabase config (diisi otomatis dari st.secrets saat di cloud) ────────────
def _get_supabase_cfg():
    """Ambil config Supabase: dari st.secrets (cloud) atau hardcode (lokal)."""
    try:
        return st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"], st.secrets["SUPABASE_BUCKET"]
    except Exception:
        return (
            "https://ccscayzxjmwzrxgdrqos.supabase.co",
            "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNjc2NheXp4am13enJ4Z2RycW9zIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODI0NjEyNDksImV4cCI6MjA5ODAzNzI0OX0.OQuopCqO28WcYFfbFvuVicj_OvHonduB_QRCrMZdnY",
            "dashboard-data",
        )


PAYMENT_SCHEDULE = {
    "KSNI":  [0, 1, 2, 3, 4],  # Senin-Jumat
    "NSI":   [0],              # Senin
    "SIMBA": [1, 3],           # Selasa & Kamis
    "MEIJI": [3],              # Kamis
    "NNI":   [0],              # Senin
}

LUNAS_VALUES   = {"lunas", "lunas "}
SIAP_BAYAR_VAL = {"siap bayar"}

# ──────────────────────────────────────────────────────────────────────────────
#  LOGGING
# ──────────────────────────────────────────────────────────────────────────────

LOG_DIR.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_DIR / "dashboard.log", encoding="utf-8")],
)

# ──────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ──────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Principal Payment Planning",
    page_icon="🔴",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────────────────────────────────────
#  CSS
# ──────────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    color: #212529;
}
.main .block-container { padding: 0 1.5rem 2rem 1.5rem; max-width: 1400px; }
#MainMenu, footer, .stDeployButton { display: none !important; }

/* ── Header ── */
.top-header {
    display: flex; align-items: center; justify-content: space-between;
    padding: 0.8rem 1.4rem;
    background: #fff;
    border-bottom: 2px solid #C0392B;
    margin-bottom: 1.25rem;
    box-shadow: 0 1px 4px rgba(0,0,0,.06);
}
.top-header img { height: 36px; object-fit: contain; }
.brand-title   { font-size: .95rem; font-weight: 700; color: #C0392B; line-height: 1.2; }
.brand-sub     { font-size: .68rem; color: #6C757D; }
.badge-ok      { background:#D4EDDA; color:#155724; border:1px solid #C3E6CB;
                 padding:.2rem .6rem; border-radius:10px; font-size:.7rem; font-weight:600; }
.badge-wait    { background:#FFF3CD; color:#856404; border:1px solid #FFEEBA;
                 padding:.2rem .6rem; border-radius:10px; font-size:.7rem; font-weight:600; }

/* ── Section title ── */
.sec-title {
    font-size: .75rem; font-weight: 700; color: #6C757D;
    text-transform: uppercase; letter-spacing: .06em;
    border-left: 3px solid #C0392B;
    padding-left: .5rem;
    margin: 1.25rem 0 .6rem 0;
}

/* ── KPI card ── */
.kpi-card {
    background: #fff; border: 1px solid #DEE2E6;
    border-radius: 7px; padding: .85rem 1rem;
    box-shadow: 0 1px 3px rgba(0,0,0,.04);
}
.kpi-card.hi { border-left: 3px solid #C0392B; }
.kpi-lbl  { font-size:.68rem; font-weight:600; color:#6C757D;
             text-transform:uppercase; letter-spacing:.05em; margin-bottom:.2rem; }
.kpi-val  { font-size:1.35rem; font-weight:700; color:#212529; }
.kpi-val.red { color:#C0392B; }
.kpi-sub  { font-size:.65rem; color:#ADB5BD; margin-top:.15rem; }

/* ── Status pill ── */
.pill-row { display:flex; flex-wrap:wrap; gap:.5rem; margin-bottom:.75rem; }

/* ── Table ── */
.dataframe { font-size:.77rem !important; }

/* ── Schedule table ── */
.sched-tbl { width:100%; border-collapse:collapse; font-size:.79rem; }
.sched-tbl th { background:#F8F9FA; color:#6C757D; font-size:.67rem;
                font-weight:600; text-transform:uppercase; letter-spacing:.04em;
                padding:.45rem .7rem; border-bottom:1px solid #DEE2E6; text-align:left; }
.sched-tbl td { padding:.45rem .7rem; border-bottom:1px solid #DEE2E6; vertical-align:middle; }
.sched-tbl tr:last-child td { border-bottom:none; }
.tag-today { background:#FADBD8; color:#C0392B; padding:.15rem .45rem;
             border-radius:4px; font-size:.67rem; font-weight:700; }
.tag-soon  { background:#FFF3CD; color:#856404; padding:.15rem .45rem;
             border-radius:4px; font-size:.67rem; font-weight:700; }
.tag-later { background:#F8F9FA; color:#6C757D; padding:.15rem .45rem;
             border-radius:4px; font-size:.67rem; }

/* ── Sidebar ── */
[data-testid="stSidebar"] { background:#F8F9FA; border-right:1px solid #DEE2E6; }
.stDownloadButton > button {
    background:#F8F9FA; border:1px solid #DEE2E6; color:#212529;
    font-size:.76rem; border-radius:5px; padding:.35rem .8rem;
}
.stDownloadButton > button:hover { border-color:#C0392B; color:#C0392B; background:#FADBD8; }
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
#  HELPER FUNCTIONS
# ──────────────────────────────────────────────────────────────────────────────

def fmt_rp(v, short=True):
    if pd.isna(v) or v == 0: return "Rp 0"
    sign = "-" if v < 0 else ""
    a = abs(v)
    if short:
        if a >= 1e12: return f"{sign}Rp {a/1e12:.2f} T"
        if a >= 1e9:  return f"{sign}Rp {a/1e9:.1f} M"
        if a >= 1e6:  return f"{sign}Rp {a/1e6:.1f} Jt"
    return f"{sign}Rp {a:,.0f}"

def next_schedule(from_date=None):
    if from_date is None:
        from_date = date.today()
    result = {}
    for p, days in PAYMENT_SCHEDULE.items():
        if set(days) == set(range(7)):
            result[p] = from_date; continue
        for i in range(8):
            d = from_date + timedelta(days=i)
            if d.weekday() in days:
                result[p] = d; break
    return result

def _supabase_get(filename: str) -> bytes | None:
    """Download file dari Supabase Storage."""
    url, key, bucket = _get_supabase_cfg()
    try:
        # Coba via authenticated endpoint dulu
        resp = requests.get(
            f"{url}/storage/v1/object/{bucket}/{filename}",
            headers={"Authorization": f"Bearer {key}"},
            timeout=30,
        )
        if resp.status_code == 200:
            return resp.content
        # Kalau gagal, coba public endpoint
        resp2 = requests.get(
            f"{url}/storage/v1/object/public/{bucket}/{filename}",
            timeout=30,
        )
        if resp2.status_code == 200:
            return resp2.content
        # Log detail error untuk debug
        st.sidebar.error(f"Supabase error {resp.status_code}: {resp.text[:200]}")
        return None
    except Exception as e:
        st.sidebar.error(f"Supabase exception: {e}")
        return None


def load_meta():
    # Coba dari Supabase dulu
    data = _supabase_get("last_update.json")
    if data:
        try:
            return json.loads(data)
        except Exception:
            pass
    # Fallback lokal
    if LAST_UPDATE_FILE.exists():
        try:
            return json.loads(LAST_UPDATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


@st.cache_data(ttl=300, show_spinner=False)
def load_data():
    # Coba dari Supabase dulu
    data = _supabase_get("Dashboard_Data.xlsx")
    source = "Supabase"

    # Fallback ke lokal
    if data is None:
        if not DASHBOARD_DATA.exists():
            return None
        data = DASHBOARD_DATA.read_bytes()
        source = "lokal"

    try:
        df = pd.read_excel(
            io.BytesIO(data),
            sheet_name="Invoice",
            engine="openpyxl",
            dtype={"No Invoice": str, "Principal": str,
                   "Vendor": str, "PIC": str,
                   "Status ACC": str, "Kategori": str, "Area": str},
        )
        logging.info("Data dimuat dari %s: %d baris", source, len(df))
        # Tanggal
        for col in ["Tgl Invoice", "Tgl Jatuh Tempo", "Tgl Payment"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        # Numerik
        for col in ["Nominal Invoice", "Nominal Bayar", "Sisa Tagihan", "Aging (Hari)"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        # Flag
        if "Kategori" in df.columns:
            norm = df["Kategori"].astype(str).str.strip().str.lower()
            df["_lunas"]      = norm.isin(LUNAS_VALUES)
            df["_siap_bayar"] = norm.isin(SIAP_BAYAR_VAL)
        return df
    except Exception as e:
        st.error(f"Gagal baca data: {e}")
        return None

def logo_html():
    # Coba beberapa lokasi: cloud path, lokal, fallback teks
    for path in [LOGO_PATH_CLOUD, LOGO_PATH]:
        if path.exists():
            try:
                b64 = base64.b64encode(path.read_bytes()).decode()
                ext = "jpeg" if path.suffix.lower() in (".jpg",".jpeg") else path.suffix.lstrip(".")
                return f'<img src="data:image/{ext};base64,{b64}" alt="logo">'
            except Exception:
                pass
    return '<span style="font-size:1.4rem;font-weight:800;color:#C0392B;">PMA</span>'


def chart_cfg():
    return {"displayModeBar": False, "responsive": True}

BASE_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter,sans-serif", color="#212529", size=11),
    margin=dict(l=6, r=6, t=32, b=6),
    colorway=["#C0392B","#2980B9","#27AE60","#F39C12","#8E44AD","#16A085"],
)


# ──────────────────────────────────────────────────────────────────────────────
#  HEADER
# ──────────────────────────────────────────────────────────────────────────────

def render_header(meta):
    updated = False
    date_str = meta.get("last_update_date", "—")
    time_str = meta.get("last_update_time", "—")
    try:
        updated = datetime.strptime(date_str, "%Y-%m-%d").date() == date.today()
    except Exception:
        pass
    badge = ('<span class="badge-ok">● Data Updated</span>' if updated
             else '<span class="badge-wait">● Menunggu Update Hari Ini</span>')
    st.markdown(f"""
    <div class="top-header">
        <div style="display:flex;align-items:center;gap:.7rem;">
            {logo_html()}
            <div>
                <div class="brand-title">Principal Payment Planning Dashboard</div>
                <div class="brand-sub">PT Pinus Merah Abadi &nbsp;·&nbsp; Finance Account Payable</div>
            </div>
        </div>
        <div style="text-align:right;">
            <div style="font-size:.68rem;color:#6C757D;margin-bottom:.25rem;">
                Last Update &nbsp;<b style="color:#212529;">{date_str}</b>&nbsp;{time_str}
                &nbsp;|&nbsp; Dashboard_Data.xlsx
            </div>
            {badge}
        </div>
    </div>
    """, unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
#  KPI CARDS
# ──────────────────────────────────────────────────────────────────────────────

# Mapping kategori → grup display
KATEGORI_GRUP = {
    "belum miro":  "Belum Miro",
    "proses pa":   "Proses PA",
    "siap buat pa":"Proses PA",
    "siap bayar":  "Siap Bayar",
    "lunas":       "Lunas",
    "lunas ":      "Lunas",
}

def render_kpi(df):
    st.markdown('<div class="sec-title">Executive Summary</div>', unsafe_allow_html=True)

    lunas   = df["_lunas"]      if "_lunas"      in df.columns else pd.Series(False, index=df.index)
    siap    = df["_siap_bayar"] if "_siap_bayar" in df.columns else pd.Series(False, index=df.index)
    nominal = df["Nominal Invoice"] if "Nominal Invoice" in df.columns else pd.Series(0.0, index=df.index)

    outstanding = float(nominal[~lunas].sum())
    rtp         = float(nominal[siap].sum())
    avg6        = _avg6(df)

    # ── Baris 1: KPI Utama ───────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    _kpi(c1, "Total Outstanding", fmt_rp(outstanding), "Invoice belum Lunas", hi=True)
    _kpi(c2, "Siap Bayar",        fmt_rp(rtp),         "Menunggu pembayaran",  hi=True)
    _kpi(c3, "Jumlah Invoice",    f"{len(df):,}",       "Total invoice")
    avg_aging = df["Aging (Hari)"].mean() if "Aging (Hari)" in df.columns else 0
    _kpi(c4, "Rata-rata Aging",   f"{avg_aging:.0f} hari", "Rata-rata outstanding")

    st.markdown("<div style='height:.6rem'></div>", unsafe_allow_html=True)

    # ── Baris 2: Breakdown Kategori ──────────────────────────────────────────
    st.markdown('<div style="font-size:.7rem;font-weight:600;color:#6C757D;'
                'text-transform:uppercase;letter-spacing:.05em;margin-bottom:.4rem;">'
                'Breakdown Kategori</div>', unsafe_allow_html=True)

    # Hitung per grup
    URUTAN = ["Belum Miro", "Proses PA", "Siap Bayar", "Lunas", "Case"]
    hasil  = {g: {"nominal": 0.0, "count": 0} for g in URUTAN}

    if "Kategori" in df.columns:
        for _, row in df.iterrows():
            kat  = str(row.get("Kategori", "")).strip().lower()
            nom  = float(row.get("Nominal Invoice", 0) or 0)
            grup = KATEGORI_GRUP.get(kat, "Case")
            hasil[grup]["nominal"] += nom
            hasil[grup]["count"]   += 1

    cols = st.columns(5)
    WARNA = {
        "Belum Miro": "#2980B9",
        "Proses PA":  "#F39C12",
        "Siap Bayar": "#27AE60",
        "Lunas":      "#6C757D",
        "Case":       "#C0392B",
    }
    for col, grup in zip(cols, URUTAN):
        d   = hasil[grup]
        warna = WARNA[grup]
        col.markdown(f"""
        <div style="background:#fff;border:1px solid #DEE2E6;border-top:3px solid {warna};
                    border-radius:6px;padding:.7rem .9rem;">
            <div style="font-size:.68rem;font-weight:600;color:#6C757D;
                        text-transform:uppercase;letter-spacing:.04em;margin-bottom:.3rem;">
                {grup}
            </div>
            <div style="font-size:1.05rem;font-weight:700;color:{warna};">
                {fmt_rp(d["nominal"])}
            </div>
            <div style="font-size:.72rem;color:#ADB5BD;margin-top:.15rem;">
                {d["count"]:,} invoice
            </div>
        </div>""", unsafe_allow_html=True)

    return {"outstanding": outstanding, "rtp": rtp, "forecast": 0, "backlog": 0, "avg6": avg6}

def _kpi(col, label, value, sub, hi=False):
    cls = "kpi-card hi" if hi else "kpi-card"
    vcls = "kpi-val red" if hi else "kpi-val"
    col.markdown(f"""
    <div class="{cls}">
        <div class="kpi-lbl">{label}</div>
        <div class="{vcls}">{value}</div>
        <div class="kpi-sub">{sub}</div>
    </div>""", unsafe_allow_html=True)

def _avg6(df):
    """Rata-rata nominal bayar per bulan, 6 bulan terakhir."""
    lunas = df.get("_lunas", pd.Series(False, index=df.index))
    if "Tgl Payment" in df.columns and "Nominal Bayar" in df.columns:
        src = df[lunas & (df["Nominal Bayar"] > 0)].copy()
        if not src.empty:
            src["_p"] = src["Tgl Payment"].dt.to_period("M")
            monthly = src.groupby("_p")["Nominal Bayar"].sum().sort_index().tail(6)
            if len(monthly): return float(monthly.mean())
    # Fallback via Nominal Invoice
    if "Nominal Invoice" in df.columns and "Bulan" in df.columns and "Tahun" in df.columns:
        src = df[lunas].copy()
        if not src.empty:
            monthly = src.groupby(["Tahun","Bulan"])["Nominal Invoice"].sum().tail(6)
            if len(monthly): return float(monthly.mean())
    return 0.0


# ──────────────────────────────────────────────────────────────────────────────
#  STATUS / KATEGORI PILLS
# ──────────────────────────────────────────────────────────────────────────────

def render_kategori_pills(df):
    st.markdown('<div class="sec-title">Distribusi Kategori</div>', unsafe_allow_html=True)
    if "Kategori" not in df.columns:
        st.info("Kolom Kategori tidak ditemukan.")
        return

    dist = (
        df.groupby("Kategori")
        .agg(Count=("No Invoice","count"),
             Amount=("Nominal Invoice","sum"))
        .reset_index()
        .sort_values("Amount", ascending=False)
    )

    selected = st.session_state.get("sel_kat", None)
    cols = st.columns(min(len(dist), 5))
    for i, row in dist.iterrows():
        col = cols[i % len(cols)]
        is_active = selected == row["Kategori"]
        border = "2px solid #C0392B" if is_active else "1px solid #DEE2E6"
        bg     = "#FADBD8" if is_active else "#F8F9FA"
        if col.button(
            f"{row['Kategori']}\n{int(row['Count'])} invoice\n{fmt_rp(row['Amount'])}",
            key=f"pill_{row['Kategori']}",
            use_container_width=True
        ):
            st.session_state["sel_kat"] = None if is_active else row["Kategori"]
            st.rerun()


# ──────────────────────────────────────────────────────────────────────────────
#  CHARTS
# ──────────────────────────────────────────────────────────────────────────────

def chart_trend(df):
    lunas = df.get("_lunas", pd.Series(False, index=df.index))
    src   = df[lunas].copy()
    if "Tgl Payment" in src.columns and "Nominal Bayar" in src.columns:
        src = src[src["Nominal Bayar"] > 0]
        if not src.empty:
            src["Period"] = src["Tgl Payment"].dt.to_period("M")
            trend = src.groupby("Period")["Nominal Bayar"].sum().sort_index().tail(12).reset_index()
            trend["Period"] = trend["Period"].astype(str)
            trend["MA6"]    = trend["Nominal Bayar"].rolling(6, min_periods=1).mean()

            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=trend["Period"], y=trend["Nominal Bayar"],
                name="Nominal Bayar",
                marker_color="#FADBD8", marker_line_color="#C0392B", marker_line_width=1,
                hovertemplate="<b>%{x}</b><br>%{customdata}<extra></extra>",
                customdata=[fmt_rp(v) for v in trend["Nominal Bayar"]],
            ))
            fig.add_trace(go.Scatter(
                x=trend["Period"], y=trend["MA6"],
                name="Rata-rata 6 Bln", mode="lines+markers",
                line=dict(color="#C0392B", width=2, dash="dash"),
                marker=dict(size=5),
                hovertemplate="<b>%{x}</b><br>Avg: %{customdata}<extra></extra>",
                customdata=[fmt_rp(v) for v in trend["MA6"]],
            ))
            fig.update_layout(
                title=dict(text="Trend Pembayaran Bulanan", font=dict(size=12)),
                xaxis=dict(tickfont=dict(size=10), gridcolor="#DEE2E6"),
                yaxis=dict(tickfont=dict(size=10), gridcolor="#DEE2E6"),
                bargap=0.35, **BASE_LAYOUT
            )
            return fig
    return None

def chart_kategori_bar(df):
    if "Kategori" not in df.columns: return None
    dist = (df.groupby("Kategori")["Nominal Invoice"]
            .sum().sort_values().reset_index())
    colors = ["#C0392B" if "Lunas" in k else
              "#27AE60" if "Siap" in k else "#2980B9"
              for k in dist["Kategori"]]
    fig = go.Figure(go.Bar(
        x=dist["Nominal Invoice"], y=dist["Kategori"],
        orientation="h", marker_color=colors,
        text=[fmt_rp(v) for v in dist["Nominal Invoice"]],
        textposition="outside",
        hovertemplate="<b>%{y}</b><br>%{text}<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text="Nominal per Kategori", font=dict(size=12)),
        xaxis=dict(tickfont=dict(size=9), gridcolor="#DEE2E6"),
        yaxis=dict(tickfont=dict(size=10), gridcolor="rgba(0,0,0,0)"),
        **BASE_LAYOUT
    )
    fig.update_layout(margin=dict(l=10, r=80, t=32, b=6))
    return fig

def chart_aging(df):
    if "Aging Bucket" not in df.columns: return None
    dist = df.groupby("Aging Bucket").agg(
        Count=("No Invoice","count"),
        Amount=("Nominal Invoice","sum")
    ).reset_index()
    bucket_colors = {
        "0-7 Hari": "#27AE60", "8-14 Hari": "#F39C12",
        "15-30 Hari": "#E67E22", "31+ Hari": "#C0392B",
    }
    colors = [bucket_colors.get(b, "#2980B9") for b in dist["Aging Bucket"]]
    fig = go.Figure(go.Bar(
        x=dist["Aging Bucket"], y=dist["Count"],
        marker_color=colors,
        text=dist["Count"].astype(str), textposition="outside",
        hovertemplate="<b>%{x}</b><br>Count: %{y}<br>%{customdata}<extra></extra>",
        customdata=[fmt_rp(v) for v in dist["Amount"]],
    ))
    fig.update_layout(
        title=dict(text="Distribusi Aging", font=dict(size=12)),
        xaxis=dict(tickfont=dict(size=10), gridcolor="rgba(0,0,0,0)"),
        yaxis=dict(tickfont=dict(size=10), gridcolor="#DEE2E6"),
        bargap=0.4, **BASE_LAYOUT
    )
    return fig

def chart_principal(df):
    if "Principal" not in df.columns: return None
    lunas = df.get("_lunas", pd.Series(False, index=df.index))
    grp = (df[~lunas].groupby("Principal")["Nominal Invoice"]
           .sum().sort_values().reset_index())
    fig = go.Figure(go.Bar(
        x=grp["Nominal Invoice"], y=grp["Principal"],
        orientation="h", marker_color="#C0392B", marker_opacity=0.8,
        text=[fmt_rp(v) for v in grp["Nominal Invoice"]], textposition="outside",
        hovertemplate="<b>%{y}</b><br>%{text}<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text="Outstanding per Principal", font=dict(size=12)),
        xaxis=dict(tickfont=dict(size=9), gridcolor="#DEE2E6"),
        yaxis=dict(tickfont=dict(size=10), gridcolor="rgba(0,0,0,0)"),
        **BASE_LAYOUT
    )
    fig.update_layout(margin=dict(l=10, r=80, t=32, b=6))
    return fig

def render_chart(fig, height=290, key=""):
    if fig is None:
        st.markdown(
            f"<div style='height:{height}px;display:flex;align-items:center;"
            "justify-content:center;color:#ADB5BD;font-size:.8rem;"
            "background:#F8F9FA;border:1px dashed #DEE2E6;border-radius:6px;'>"
            "Tidak ada data</div>", unsafe_allow_html=True)
    else:
        st.plotly_chart(fig, use_container_width=True, config=chart_cfg(), key=key or None)


# ──────────────────────────────────────────────────────────────────────────────
#  PAYMENT SCHEDULE
# ──────────────────────────────────────────────────────────────────────────────

# Berapa kali principal bayar per minggu
FREKUENSI_BAYAR = {
    "KSNI":  5,   # Senin-Jumat
    "NSI":   1,   # 1x seminggu (Senin)
    "SIMBA": 2,   # 2x seminggu (Selasa & Kamis)
    "MEIJI": 1,   # 1x seminggu (Kamis)
    "NNI":   1,   # 1x seminggu (Senin)
}

def _forecast_per_principal(df_raw, principal, next_pay_date):
    """
    Forecast payment untuk tanggal next_pay_date (satu sesi bayar).
    Logika:
      - Hitung total bayar 6 bulan terakhir per principal
      - Bagi dengan total sesi bayar dalam 6 bulan sesuai frekuensi
        (mis. KSNI: 26*7=182 hari, SIMBA: 26*2=52 sesi)
      - Hasilnya = rata-rata per SESI bayar
      - Bandingkan dengan Siap Bayar → ambil MIN
    """
    if df_raw is None or df_raw.empty:
        return 0.0

    p_df = df_raw[df_raw["Principal"] == principal] if "Principal" in df_raw.columns else df_raw
    lunas = p_df.get("_lunas", pd.Series(False, index=p_df.index))

    # Frekuensi sesi bayar per minggu
    freq = FREKUENSI_BAYAR.get(principal, 1)
    total_sesi = 26 * freq  # 6 bulan = 26 minggu

    # Total bayar 6 bulan terakhir
    total_paid = 0.0
    if "Tgl Payment" in p_df.columns and "Nominal Bayar" in p_df.columns:
        paid = p_df[lunas & (p_df["Nominal Bayar"] > 0)].copy()
        if not paid.empty:
            cutoff = pd.Timestamp(date.today()) - pd.DateOffset(months=6)
            paid = paid[paid["Tgl Payment"] >= cutoff]
            total_paid = float(paid["Nominal Bayar"].sum())

    # Rata-rata per sesi
    avg_per_sesi = total_paid / total_sesi if total_sesi > 0 else 0.0

    # Siap Bayar s.d. tanggal bayar berikutnya
    siap = 0.0
    siap_mask = p_df.get("_siap_bayar", pd.Series(False, index=p_df.index))
    if "Tgl Jatuh Tempo" in p_df.columns and "Nominal Invoice" in p_df.columns:
        due = pd.Timestamp(next_pay_date)
        siap = float(p_df.loc[siap_mask & (p_df["Tgl Jatuh Tempo"] <= due), "Nominal Invoice"].sum())
        if siap == 0:
            siap = float(p_df.loc[siap_mask, "Nominal Invoice"].sum())

    if avg_per_sesi <= 0 and siap <= 0:
        return 0.0
    if avg_per_sesi <= 0:
        return siap
    return min(avg_per_sesi, siap) if siap > 0 else avg_per_sesi

def render_schedule(df_raw=None):
    st.markdown('<div class="sec-title">Jadwal Pembayaran</div>', unsafe_allow_html=True)
    today    = date.today()
    schedule = next_schedule(today)
    rows     = ""
    for p, d in schedule.items():
        delta = (d - today).days
        if delta == 0:   tag = f'<span class="tag-today">Hari Ini</span>'
        elif delta <= 3: tag = f'<span class="tag-soon">Dalam {delta} Hari</span>'
        else:            tag = f'<span class="tag-later">Dalam {delta} Hari</span>'

        # Forecast minggu ini per principal
        fcst = _forecast_per_principal(df_raw, p, d) if df_raw is not None else 0.0
        fcst_str = fmt_rp(fcst) if fcst > 0 else '<span style="color:#ADB5BD">—</span>'

        rows += (
            f"<tr>"
            f"<td><b>{p}</b></td>"
            f"<td>{d.strftime('%A, %d %B %Y')}</td>"
            f"<td>{tag}</td>"
            f"<td style='font-weight:600;color:#C0392B;'>{fcst_str}</td>"
            f"</tr>"
        )
    st.markdown(f"""
    <table class="sched-tbl">
        <thead>
            <tr>
                <th>Principal</th>
                <th>Tanggal Pembayaran</th>
                <th>Status</th>
                <th>Forecast Payment</th>
            </tr>
        </thead>
        <tbody>{rows}</tbody>
    </table>""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
#  FORECAST BOX
# ──────────────────────────────────────────────────────────────────────────────

def render_forecast_box(kpis):
    st.markdown('<div class="sec-title">Forecast Payment</div>', unsafe_allow_html=True)
    items = [
        ("Siap Bayar",              fmt_rp(kpis["rtp"])),
        ("Rata-rata Bayar 6 Bln",   fmt_rp(kpis["avg6"])),
        ("Forecast Payment",        fmt_rp(kpis["forecast"])),
        ("Backlog",                 fmt_rp(kpis["backlog"])),
    ]
    cols = st.columns(4)
    for col, (label, val) in zip(cols, items):
        col.markdown(f"""
        <div style="background:#F8F9FA;border:1px solid #DEE2E6;border-radius:6px;
                    padding:.7rem .9rem;">
            <div style="font-size:.67rem;font-weight:600;color:#6C757D;
                        text-transform:uppercase;letter-spacing:.05em;">{label}</div>
            <div style="font-size:1.1rem;font-weight:700;color:#C0392B;">{val}</div>
        </div>""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
#  INVOICE TABLE
# ──────────────────────────────────────────────────────────────────────────────

def render_tabel(df):
    st.markdown('<div class="sec-title">Detail Invoice</div>', unsafe_allow_html=True)
    if df.empty:
        st.info("Tidak ada data sesuai filter.")
        return

    SHOW_COLS = [c for c in [
        "Principal", "Vendor", "PIC",
        "No Invoice", "Tgl Invoice", "Tgl Jatuh Tempo",
        "Area", "Nominal Invoice", "DPP", "PPN",
        "No BPPR", "No BASP",
        "Tgl Miro", "No Miro",
        "Nominal Bayar", "No PA",
        "Tgl Clearing", "No Clearing",
        "Sisa Tagihan",
        "Status ACC", "Kategori", "Keterangan",
        "Aging (Hari)", "Aging Bucket",
    ] if c in df.columns]

    disp = df[SHOW_COLS].copy()
    for col in ["Tgl Invoice", "Tgl Jatuh Tempo"]:
        if col in disp.columns:
            disp[col] = pd.to_datetime(disp[col], errors="coerce").dt.strftime("%d-%m-%Y")
    for col in ["Nominal Invoice", "Nominal Bayar", "Sisa Tagihan"]:
        if col in disp.columns:
            disp[col] = pd.to_numeric(disp[col], errors="coerce").apply(
                lambda x: f"Rp {x:,.0f}" if pd.notna(x) and x > 0 else "—")
    if "Aging (Hari)" in disp.columns:
        disp["Aging (Hari)"] = disp["Aging (Hari)"].apply(
            lambda x: f"{int(x)} hr" if pd.notna(x) else "—")

    # Pagination
    PAGE_SIZE   = 50
    total       = len(disp)
    total_pages = max(1, (total - 1) // PAGE_SIZE + 1)

    c1, _, c3 = st.columns([2, 3, 2])
    c1.caption(f"{total:,} invoice")
    page = c3.number_input("Halaman", min_value=1, max_value=total_pages,
                           value=1, step=1, label_visibility="collapsed")
    c3.caption(f"Hal. {page}/{total_pages}")

    start = (page - 1) * PAGE_SIZE
    st.dataframe(disp.iloc[start:start + PAGE_SIZE],
                 use_container_width=True, hide_index=True)


# ──────────────────────────────────────────────────────────────────────────────
#  DOWNLOAD
# ──────────────────────────────────────────────────────────────────────────────

def render_download(df, kpis):
    st.markdown('<div class="sec-title">Download Laporan</div>', unsafe_allow_html=True)
    st.caption("Semua download mengikuti filter aktif.")

    import io
    from openpyxl.styles import Font, PatternFill, Alignment

    def to_excel(frames_sheets: list):
        """frames_sheets = [(df, sheet_name), ...]"""
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            for frame, name in frames_sheets:
                frame.to_excel(w, sheet_name=name, index=False)
                ws = w.sheets[name]
                hf = Font(bold=True, color="FFFFFF", size=10)
                hfill = PatternFill("solid", fgColor="C0392B")
                for cell in ws[1]:
                    cell.font, cell.fill = hf, hfill
                    cell.alignment = Alignment(horizontal="center")
                ws.freeze_panes = "A2"
        buf.seek(0)
        return buf.read()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Siapkan data per laporan
    SHOW = [c for c in [
        "Principal","Vendor","PIC",
        "No Invoice","Tgl Invoice","Tgl Jatuh Tempo",
        "Area","Nominal Invoice","DPP","PPN",
        "No BPPR","No BASP",
        "Tgl Miro","No Miro",
        "Nominal Bayar","No PA",
        "Tgl Clearing","No Clearing",
        "Sisa Tagihan",
        "Status ACC","Kategori","Keterangan",
        "Aging (Hari)","Aging Bucket",
    ] if c in df.columns]
    detail_df = df[SHOW].copy()
    for col in ["Tgl Invoice","Tgl Jatuh Tempo"]:
        if col in detail_df.columns:
            detail_df[col] = pd.to_datetime(detail_df[col], errors="coerce").dt.strftime("%d-%m-%Y")

    lunas = df.get("_lunas", pd.Series(False, index=df.index))
    siap  = df.get("_siap_bayar", pd.Series(False, index=df.index))

    # Principal summary
    def _row(x):
        return pd.Series({
            "Total Invoice":      len(x),
            "Lunas":              int(x["_lunas"].sum()),
            "Siap Bayar":         int(x["_siap_bayar"].sum()),
            "Outstanding":        float(x.loc[~x["_lunas"],"Nominal Invoice"].sum()),
            "Nominal Siap Bayar": float(x.loc[x["_siap_bayar"],"Nominal Invoice"].sum()),
        })
    prin_sum = df.groupby("Principal").apply(_row, include_groups=False).reset_index() \
               if "Principal" in df.columns else pd.DataFrame()

    # Kategori summary
    kat_sum = (df.groupby("Kategori")
               .agg(Jumlah=("No Invoice","count"), Total_Nominal=("Nominal Invoice","sum"))
               .reset_index()
               .sort_values("Total_Nominal", ascending=False)
               ) if "Kategori" in df.columns else pd.DataFrame()

    # Forecast
    forecast_df = pd.DataFrame([
        {"Item": "Total Outstanding",    "Nominal (Rp)": kpis["outstanding"]},
        {"Item": "Siap Bayar",           "Nominal (Rp)": kpis["rtp"]},
        {"Item": "Rata-rata Bayar 6Bln", "Nominal (Rp)": kpis["avg6"]},
        {"Item": "Forecast Payment",     "Nominal (Rp)": kpis["forecast"]},
        {"Item": "Backlog",              "Nominal (Rp)": kpis["backlog"]},
    ])

    c1,c2,c3,c4,c5 = st.columns(5)
    with c1:
        st.download_button("📥 Invoice Detail (.xlsx)",
            to_excel([(detail_df, "Invoice Detail")]),
            f"invoice_detail_{ts}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with c2:
        st.download_button("📥 Invoice Detail (.csv)",
            detail_df.to_csv(index=False).encode("utf-8"),
            f"invoice_detail_{ts}.csv", "text/csv",
            use_container_width=True)
    with c3:
        st.download_button("📥 Ringkasan Principal",
            to_excel([(prin_sum, "Principal")]),
            f"principal_{ts}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with c4:
        st.download_button("📥 Ringkasan Kategori",
            to_excel([(kat_sum, "Kategori")]),
            f"kategori_{ts}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with c5:
        st.download_button("📥 Forecast Report",
            to_excel([(forecast_df, "Forecast")]),
            f"forecast_{ts}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)


# ──────────────────────────────────────────────────────────────────────────────
#  SIDEBAR FILTER
# ──────────────────────────────────────────────────────────────────────────────

def render_sidebar(df):
    with st.sidebar:
        st.markdown("**Filter**")

        def opts(col):
            if col not in df.columns: return []
            return sorted(df[col].dropna().astype(str).unique().tolist())

        search    = st.text_input("Cari No Invoice / Keterangan", placeholder="Ketik...")
        principal = st.multiselect("Principal", opts("Principal"))
        pic       = st.multiselect("PIC",       opts("PIC"))
        kategori  = st.multiselect("Kategori",  opts("Kategori"))
        area      = st.multiselect("Area",      opts("Area"))

        st.markdown("---")
        if st.button("Reset Filter", use_container_width=True):
            st.session_state["sel_kat"] = None
            st.rerun()

    return dict(search=search, principal=principal, pic=pic,
                kategori=kategori, area=area)


def apply_filter(df, f):
    if f["principal"]: df = df[df["Principal"].isin(f["principal"])]
    if f["pic"]      : df = df[df["PIC"].isin(f["pic"])]
    if f["kategori"] and "Kategori" in df.columns:
        df = df[df["Kategori"].isin(f["kategori"])]
    if f["area"] and "Area" in df.columns:
        df = df[df["Area"].isin(f["area"])]
    if f["search"]:
        q = f["search"].strip().lower()
        mask = pd.Series(False, index=df.index)
        for col in ["No Invoice","Principal","Keterangan","Area"]:
            if col in df.columns:
                mask |= df[col].astype(str).str.lower().str.contains(q, na=False)
        df = df[mask]
    return df


# ──────────────────────────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=300, show_spinner=False)
def load_retur():
    """Load sheet Retur dari Dashboard_Data.xlsx via Supabase."""
    data = _supabase_get("Dashboard_Data.xlsx")
    if data is None:
        if not DASHBOARD_DATA.exists():
            return None, None
        data = DASHBOARD_DATA.read_bytes()
    try:
        df_retur = pd.read_excel(io.BytesIO(data), sheet_name="Retur",
                                 engine="openpyxl", dtype=str)
        df_saran = pd.read_excel(io.BytesIO(data), sheet_name="Saran PA Retur",
                                 engine="openpyxl", dtype=str)
        # Numerik
        for df in [df_retur, df_saran]:
            for col in ["Nominal Retur","DPP","PPN","Sisa Tagihan",
                        "Nilai Invoice PA","Sisa Setelah Dipotong"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        # Tanggal
        for df in [df_retur, df_saran]:
            for col in ["Tgl CN","Tgl Jatuh Tempo","Tgl Miro","Tgl Payment","Tgl Clearing"]:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
        # Flag
        for df in [df_retur, df_saran]:
            if "Kategori" in df.columns:
                norm = df["Kategori"].astype(str).str.strip().str.lower()
                df["_lunas"] = norm.isin({"lunas","lunas "})
        return df_retur, df_saran
    except Exception as e:
        return None, None


def main():
    meta = load_meta()
    render_header(meta)

    df_raw = load_data()

    # Tab navigasi
    tab1, tab2 = st.tabs(["📄 Outstanding Invoice", "🔄 Outstanding Retur"])

    with tab1:
        _page_invoice(df_raw)

    with tab2:
        _page_retur()


def _page_invoice(df_raw):
    # Sidebar
    filters = render_sidebar(df_raw) if df_raw is not None else {}

    if df_raw is None or df_raw.empty:
        st.markdown("""
        <div style="text-align:center;padding:4rem 2rem;background:#F8F9FA;
                    border:1px dashed #DEE2E6;border-radius:8px;margin:2rem 0;">
            <div style="font-size:2.5rem;">📂</div>
            <div style="font-size:1rem;font-weight:600;margin:.5rem 0;">
                Dashboard_Data.xlsx Belum Tersedia</div>
            <div style="font-size:.82rem;color:#6C757D;">
                Jalankan <b>update.exe</b> untuk generate data terbaru.</div>
        </div>""", unsafe_allow_html=True)
        render_schedule(None)
        return

    df = apply_filter(df_raw, filters)

    # ── KPI ──────────────────────────────────────────────────────────────────
    kpis = render_kpi(df)

    # ── Charts ───────────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">Analitik</div>', unsafe_allow_html=True)
    r1c1, r1c2 = st.columns([3, 2], gap="medium")
    with r1c1: render_chart(chart_trend(df), key="trend")
    with r1c2: render_chart(chart_kategori_bar(df), key="kat_bar")

    r2c1, r2c2 = st.columns(2, gap="medium")
    with r2c1: render_chart(chart_aging(df), key="aging")
    with r2c2: render_chart(chart_principal(df), key="prin")

    # ── Schedule dengan forecast per principal ───────────────────────────────
    render_schedule(df_raw)

    # ── Detail table ─────────────────────────────────────────────────────────
    render_tabel(df)

    # ── Download ─────────────────────────────────────────────────────────────
    render_download(df, kpis)


def _page_retur():
    st.markdown('<div class="sec-title">Outstanding Retur Pembelian</div>',
                unsafe_allow_html=True)

    df_retur, df_saran = load_retur()

    if df_retur is None or df_retur.empty:
        st.info("Data retur belum tersedia. Jalankan update.exe terlebih dahulu.")
        return

    # ── KPI retur ────────────────────────────────────────────────────────────
    lunas_mask = df_retur.get("_lunas", pd.Series(False, index=df_retur.index))
    belum_lunas = df_retur[~lunas_mask]

    c1, c2, c3, c4 = st.columns(4)
    _kpi(c1, "Total Retur",
         f"{len(df_retur):,}", "Semua transaksi retur")
    _kpi(c2, "Belum Lunas",
         f"{len(belum_lunas):,}", "Masih outstanding", hi=True)
    nom_col = "Nominal Retur"
    total_outstanding_retur = float(belum_lunas[nom_col].sum()) if nom_col in belum_lunas.columns else 0
    _kpi(c3, "Nominal Outstanding",
         fmt_rp(total_outstanding_retur), "Total nilai retur belum lunas", hi=True)
    sisa_col = "Sisa Tagihan"
    total_sisa = float(belum_lunas[sisa_col].sum()) if sisa_col in belum_lunas.columns else 0
    _kpi(c4, "Sisa Tagihan",
         fmt_rp(total_sisa), "Setelah pemotongan parsial")

    st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

    # ── Filter sidebar retur ──────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("---")
        st.markdown("**Filter Retur**")
        def opts_r(col):
            if col not in df_retur.columns: return []
            return sorted(df_retur[col].dropna().astype(str).unique().tolist())
        f_principal = st.multiselect("Principal (Retur)", opts_r("Principal"), key="r_principal")
        f_kategori  = st.multiselect("Kategori (Retur)",  opts_r("Kategori"),  key="r_kat")
        f_area      = st.multiselect("Area (Retur)",      opts_r("Area"),      key="r_area")
        show_lunas  = st.checkbox("Tampilkan yang Lunas", value=False, key="r_lunas")

    # Apply filter
    fdf = df_retur.copy()
    if not show_lunas:
        fdf = fdf[~fdf.get("_lunas", pd.Series(False, index=fdf.index))]
    if f_principal: fdf = fdf[fdf["Principal"].isin(f_principal)]
    if f_kategori and "Kategori" in fdf.columns:
        fdf = fdf[fdf["Kategori"].isin(f_kategori)]
    if f_area and "Area" in fdf.columns:
        fdf = fdf[fdf["Area"].isin(f_area)]

    # ── Breakdown kategori retur ──────────────────────────────────────────────
    if "Kategori" in fdf.columns:
        st.markdown('<div class="sec-title">Breakdown Kategori Retur</div>',
                    unsafe_allow_html=True)
        kat_dist = (fdf.groupby("Kategori")
                    .agg(Count=("No Retur/CN","count"),
                         Nominal=("Nominal Retur","sum"))
                    .reset_index()
                    .sort_values("Nominal", ascending=False))
        cols = st.columns(min(len(kat_dist), 5))
        for col, (_, row) in zip(cols, kat_dist.iterrows()):
            col.markdown(f"""
            <div style="background:#fff;border:1px solid #DEE2E6;border-top:3px solid #C0392B;
                        border-radius:6px;padding:.7rem .9rem;">
                <div style="font-size:.68rem;font-weight:600;color:#6C757D;
                            text-transform:uppercase;letter-spacing:.04em;">{row['Kategori']}</div>
                <div style="font-size:1.05rem;font-weight:700;color:#C0392B;">
                    {fmt_rp(row['Nominal'])}</div>
                <div style="font-size:.72rem;color:#ADB5BD;">{int(row['Count']):,} retur</div>
            </div>""", unsafe_allow_html=True)
        st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

    # ── Tabel retur ──────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">Detail Retur</div>', unsafe_allow_html=True)

    SHOW_RETUR = [c for c in [
        "Principal", "Vendor", "PIC", "No Retur/CN", "Tgl CN", "Tgl Jatuh Tempo",
        "Area", "Channel", "Nominal Retur", "Sisa Tagihan",
        "No BASP", "No BPPR", "Tgl Miro", "No Miro",
        "No PA", "Tgl Clearing", "No Clearing",
        "Status ACC", "Kategori", "Keterangan",
    ] if c in fdf.columns]

    disp = fdf[SHOW_RETUR].copy()
    for col in ["Tgl CN","Tgl Jatuh Tempo","Tgl Miro","Tgl Payment","Tgl Clearing"]:
        if col in disp.columns:
            disp[col] = pd.to_datetime(disp[col], errors="coerce").dt.strftime("%d-%m-%Y")
    for col in ["Nominal Retur","Sisa Tagihan"]:
        if col in disp.columns:
            disp[col] = pd.to_numeric(disp[col], errors="coerce").apply(
                lambda x: f"Rp {x:,.0f}" if pd.notna(x) and x > 0 else "—")

    PAGE = 50
    total = len(disp)
    total_pages = max(1, (total-1)//PAGE+1)
    c1, _, c3 = st.columns([2,3,2])
    c1.caption(f"{total:,} retur")
    page = c3.number_input("Halaman", min_value=1, max_value=total_pages,
                           value=1, step=1, label_visibility="collapsed", key="r_page")
    c3.caption(f"Hal. {page}/{total_pages}")
    start = (page-1)*PAGE
    st.dataframe(disp.iloc[start:start+PAGE], use_container_width=True, hide_index=True)

    # ── Saran PA ─────────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">Saran Matching PA untuk Retur Belum Lunas</div>',
                unsafe_allow_html=True)
    st.markdown("""
    <div style="background:#EBF5FB;border-left:3px solid #2980B9;border-radius:0 4px 4px 0;
                padding:.6rem .85rem;font-size:.79rem;color:#1A5276;margin-bottom:.75rem;">
        Saran PA dipilih dari invoice outstanding principal yang sama dengan
        <b>Nilai Invoice ≥ Nominal Retur</b>, diambil yang terkecil agar efisien.
    </div>""", unsafe_allow_html=True)

    if df_saran is not None and not df_saran.empty:
        # Filter sesuai sidebar
        fsaran = df_saran.copy()
        if f_principal: fsaran = fsaran[fsaran["Principal"].isin(f_principal)]
        if f_area and "Area" in fsaran.columns:
            fsaran = fsaran[fsaran["Area"].isin(f_area)]

        SHOW_SARAN = [c for c in [
            "Principal","Vendor","No Retur/CN","Area","Nominal Retur",
            "Sisa Tagihan","Kategori","Saran No PA","Nilai Invoice PA","Sisa Setelah Dipotong",
        ] if c in fsaran.columns]
        disp_s = fsaran[SHOW_SARAN].copy()
        for col in ["Nominal Retur","Sisa Tagihan","Nilai Invoice PA","Sisa Setelah Dipotong"]:
            if col in disp_s.columns:
                disp_s[col] = pd.to_numeric(disp_s[col], errors="coerce").apply(
                    lambda x: f"Rp {x:,.0f}" if pd.notna(x) and x > 0 else "—")
        st.dataframe(disp_s, use_container_width=True, hide_index=True)
    else:
        st.info("Saran PA belum tersedia.")

    # ── Download ─────────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">Download</div>', unsafe_allow_html=True)
    import io as _io
    from openpyxl.styles import Font, PatternFill, Alignment

    def _to_excel_retur(frames_sheets):
        buf = _io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            for frame, name in frames_sheets:
                frame.to_excel(w, sheet_name=name, index=False)
                ws = w.sheets[name]
                hf = Font(bold=True, color="FFFFFF", size=10)
                hfill = PatternFill("solid", fgColor="C0392B")
                for cell in ws[1]:
                    cell.font, cell.fill = hf, hfill
                    cell.alignment = Alignment(horizontal="center")
                ws.freeze_panes = "A2"
        buf.seek(0)
        return buf.read()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "📥 Detail Retur (.xlsx)",
            _to_excel_retur([(fdf[[c for c in SHOW_RETUR if c in fdf.columns]], "Retur")]),
            f"retur_detail_{ts}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with c2:
        if df_saran is not None and not df_saran.empty:
            st.download_button(
                "📥 Saran Matching PA (.xlsx)",
                _to_excel_retur([(fsaran[[c for c in SHOW_SARAN if c in fsaran.columns]], "Saran PA")]),
                f"saran_pa_retur_{ts}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)


if __name__ == "__main__":
    main()
