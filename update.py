"""
update.py  –  PT Pinus Merah Abadi | FAD
Baca semua .xlsb di MASTER_DATA_FOLDER, merge, clean, hitung aging,
simpan ke Dashboard_Data.xlsx + last_update.json.

Jalankan: python update.py
Compile : build_exe.bat
"""

import json, logging, sys, traceback, io
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import requests
from tqdm import tqdm

# ──────────────────────────────────────────────────────────────────────────────
#  PATH & KONFIGURASI  –  edit di sini kalau lokasi berubah
# ──────────────────────────────────────────────────────────────────────────────

MASTER_DATA_FOLDER = Path(r"D:\PROJECT FAD\MONITORING PRINSIPLE\MASTER DATA")
DASHBOARD_DATA     = Path(__file__).parent / "data" / "Dashboard_Data.xlsx"
LAST_UPDATE_FILE   = Path(__file__).parent / "data" / "last_update.json"
LOG_FILE           = Path(__file__).parent / "logs"  / "update.log"

# ──────────────────────────────────────────────────────────────────────────────
#  SUPABASE CONFIG  –  ganti dengan nilai dari project kamu
# ──────────────────────────────────────────────────────────────────────────────

SUPABASE_URL      = "https://ccscayzxjmwzrxgdrqos.supabase.co"
SUPABASE_KEY      = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNjc2NheXp4am13enJ4Z2RycW9zIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MjQ2MTI0OSwiZXhwIjoyMDk4MDM3MjQ5fQ.BtWvKtb4s57laGC5c2dcbyUpWVM45-je9jkhNHmB_8s"   # dari Settings > API Keys > service_role > Reveal
SUPABASE_BUCKET   = "dashboard-data"   # nama storage bucket yang akan dibuat

# ──────────────────────────────────────────────────────────────────────────────
#  Sheet invoice + header row per principal (dideteksi dari nama file)
SHEET_CONFIG = {
    "KSNI":  ("INVOICE",              3),
    "MEIJI": ("Invoice Meiji 2026",   4),
    "NNI":   ("GANTUNGAN - INVOICE",  3),
    "NSI":   ("INVOICE NSI 2026",     3),
    "SIMBA": ("INVOICE SIMBA'26",     3),
}

LUNAS_VALUES   = {"lunas", "lunas "}
SIAP_BAYAR_VAL = {"siap bayar"}

AGING_BUCKETS = [
    (0,   7,   "0-7 Hari"),
    (8,   14,  "8-14 Hari"),
    (15,  30,  "15-30 Hari"),
    (31,  None,"31+ Hari"),
]

# Nama kolom sumber -> nama internal
COL_MAP = {
    "nama supplier/vendor":  "Vendor",
    "vendor":                "Vendor",
    "tgl terima invoice":    "Tgl Terima",
    "tanggal terima invoice":"Tgl Terima",
    "no invoice":            "No Invoice",
    "tgl invoice":           "Tgl Invoice",
    "tanggal invoice":       "Tgl Invoice",
    "tgl jth tempo":         "Tgl Jatuh Tempo",
    "tgl jatuh tempo":       "Tgl Jatuh Tempo",
    "tanggal jatuh tempo":   "Tgl Jatuh Tempo",
    "area":                  "Area",
    "channel":               "Channel",
    "gt/mt":                 "Channel",
    "zona pa":               "Zona",
    "zona":                  "Zona",
    "nominal invoice":       "Nominal Invoice",
    "nominal":               "Nominal Invoice",
    "dpp":                   "DPP",
    " ppn ":                 "PPN",
    "ppn":                   "PPN",
    "no bppr":               "No BPPR",
    "no basp":               "No BASP",
    "tgl miro":              "Tgl Miro",
    "no miro":               "No Miro",
    "no miro/ ppu":          "No Miro",
    "tgl payment":           "Tgl Payment",
    "nominal bayar":         "Nominal Bayar",
    "nominal bayar ":        "Nominal Bayar",
    "no payment advice":     "No PA",
    "tgl clearing":          "Tgl Clearing",
    "no clearing":           "No Clearing",
    "sisa tagihan":          "Sisa Tagihan",
    "keterangan":            "Keterangan",
    "ket/ lampiran":         "Keterangan",
    "status acc":            "Status ACC",
    "status":                "Status ACC",
    "kategori":              "Kategori",
    "kategori ":             "Kategori",
    "pic":                   "PIC",
}

KOLOM_DISPLAY = [
    "Principal", "Vendor", "PIC",
    "No Invoice", "Tgl Invoice", "Tgl Jatuh Tempo",
    "Area", "Channel", "Zona",
    "Nominal Invoice", "DPP", "PPN",
    "No BPPR", "No BASP",
    "Tgl Miro", "No Miro",
    "Tgl Payment", "Nominal Bayar", "No PA",
    "Tgl Clearing", "No Clearing",
    "Sisa Tagihan",
    "Status ACC", "Kategori", "Keterangan",
    "Aging (Hari)", "Aging Bucket",
    "Bulan", "Tahun",
]

# ──────────────────────────────────────────────────────────────────────────────
#  LOGGING
# ──────────────────────────────────────────────────────────────────────────────

LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    _banner()
    log.info("=" * 60)
    log.info("Update dimulai: %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("Folder: %s", MASTER_DATA_FOLDER)

    try:
        files = _cari_file()
        if not files:
            _error("Tidak ada file .xlsb ditemukan.")
            return

        df = _baca_semua(files)
        if df.empty:
            _error("Semua file tidak menghasilkan data.")
            return

        df = _clean(df)
        df = _hitung_aging(df)
        df = _flag_status(df)
        _simpan(df)
        _simpan_meta(len(df))
        _sukses(len(df))

    except Exception as e:
        log.exception("Error fatal: %s", e)
        _error(f"{e}\n\n{traceback.format_exc()}")


# ──────────────────────────────────────────────────────────────────────────────
#  1. CARI FILE
# ──────────────────────────────────────────────────────────────────────────────

def _cari_file():
    if not MASTER_DATA_FOLDER.exists():
        log.error("Folder tidak ditemukan: %s", MASTER_DATA_FOLDER)
        return []
    files = sorted(
        f for f in MASTER_DATA_FOLDER.glob("*.xlsb")
        if not f.name.startswith(("~$", "._"))
    )
    log.info("Ditemukan %d file:", len(files))
    for f in files:
        log.info("  → %s", f.name)
    return files


# ──────────────────────────────────────────────────────────────────────────────
#  2. BACA & MERGE
# ──────────────────────────────────────────────────────────────────────────────

def _baca_semua(files):
    frames = []
    for path in tqdm(files, desc="Membaca", unit="file", ncols=65):
        try:
            df = _baca_satu(path)
            if df is not None and not df.empty:
                frames.append(df)
                log.info("  ✔ %s — %d baris", path.name, len(df))
            else:
                log.warning("  ⚠ %s — kosong", path.name)
        except Exception as e:
            log.error("  ✖ %s — %s", path.name, e)

    if not frames:
        return pd.DataFrame()

    merged = pd.concat(frames, ignore_index=True)
    log.info("Total merge: %d baris dari %d file", len(merged), len(frames))
    return merged


def _baca_satu(path):
    principal = None
    for key in SHEET_CONFIG:
        if key.upper() in path.name.upper():
            principal = key
            break

    if principal is None:
        log.warning("Principal tidak dikenali: %s — skip", path.name)
        return None

    sheet, hrow = SHEET_CONFIG[principal]
    df = pd.read_excel(path, sheet_name=sheet, engine="pyxlsb",
                       header=hrow, dtype=str)

    # Normalisasi nama kolom
    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

    rename = {col: COL_MAP[col.lower().strip()]
              for col in df.columns if col.lower().strip() in COL_MAP}
    df = df.rename(columns=rename)

    # Hapus duplikat kolom (NNI/NSI: Keterangan muncul 2x)
    df = df.loc[:, ~df.columns.duplicated(keep="first")]

    # Drop baris kosong / baris total
    df = df[df["No Invoice"].notna()]
    df = df[df["No Invoice"].astype(str).str.strip().str.len() > 0]
    df = df[~df["No Invoice"].astype(str).str.upper().str.match(
        r"^(TOTAL|SUBTOTAL|GRAND)", na=False)]

    df["Principal"] = principal
    return df.reset_index(drop=True)


# ──────────────────────────────────────────────────────────────────────────────
#  3. CLEANING
# ──────────────────────────────────────────────────────────────────────────────

def _clean(df):
    log.info("Cleaning...")
    df = df.copy()

    for col in ["Vendor", "Principal", "PIC", "Status ACC",
                "Kategori", "Keterangan", "Area", "Channel", "Zona"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace(
                {"nan": None, "None": None})

    df["No Invoice"] = df["No Invoice"].astype(str).str.strip()

    for col in ["Tgl Invoice", "Tgl Jatuh Tempo", "Tgl Payment",
                "Tgl Miro", "Tgl Clearing", "Tgl Terima"]:
        if col in df.columns:
            df[col] = df[col].apply(_parse_tgl)

    for col in ["Nominal Invoice", "DPP", "PPN", "Nominal Bayar", "Sisa Tagihan"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    before = len(df)
    df = df[df["Nominal Invoice"] > 0]
    log.info("  Drop nominal=0: %d → %d", before, len(df))

    # Tidak ada dedup — semua baris ditampilkan apa adanya
    log.info("  Baris setelah cleaning: %d", len(df))

    if "Tgl Invoice" in df.columns:
        df["Bulan"] = df["Tgl Invoice"].dt.month
        df["Tahun"] = df["Tgl Invoice"].dt.year

    return df.reset_index(drop=True)


def _parse_tgl(val):
    """Serial Excel (int) atau string → pd.Timestamp."""
    if val is None:
        return pd.NaT
    if isinstance(val, float) and pd.isna(val):
        return pd.NaT
    if isinstance(val, (int, float)):
        try:
            return pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(val))
        except Exception:
            return pd.NaT
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "-", ""):
        return pd.NaT
    try:
        return pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(float(s)))
    except Exception:
        pass
    return pd.to_datetime(s, dayfirst=True, errors="coerce")


# ──────────────────────────────────────────────────────────────────────────────
#  4. AGING
# ──────────────────────────────────────────────────────────────────────────────

def _hitung_aging(df):
    log.info("Hitung aging...")
    df = df.copy()
    today = pd.Timestamp(date.today())

    ref = "Tgl Jatuh Tempo" if "Tgl Jatuh Tempo" in df.columns else "Tgl Invoice"
    if ref in df.columns:
        df["Aging (Hari)"] = (today - df[ref]).dt.days.clip(lower=0)
    else:
        df["Aging (Hari)"] = 0
    df["Aging (Hari)"] = pd.to_numeric(df["Aging (Hari)"], errors="coerce").fillna(0).astype(int)

    def bucket(d):
        for lo, hi, label in AGING_BUCKETS:
            if hi is None and d >= lo:     return label
            if hi is not None and lo <= d <= hi: return label
        return "-"

    df["Aging Bucket"] = df["Aging (Hari)"].apply(bucket)
    log.info("  Selesai: %d invoice", len(df))
    return df


# ──────────────────────────────────────────────────────────────────────────────
#  5. FLAG STATUS
# ──────────────────────────────────────────────────────────────────────────────

def _flag_status(df):
    df = df.copy()
    if "Kategori" in df.columns:
        norm = df["Kategori"].astype(str).str.strip().str.lower()
        df["_lunas"]      = norm.isin(LUNAS_VALUES)
        df["_siap_bayar"] = norm.isin(SIAP_BAYAR_VAL)
        df.loc[df["_lunas"],      "Kategori"] = "Lunas"
        df.loc[df["_siap_bayar"], "Kategori"] = "Siap Bayar"
    else:
        df["_lunas"]      = False
        df["_siap_bayar"] = False
    return df


# ──────────────────────────────────────────────────────────────────────────────
#  6. SIMPAN
# ──────────────────────────────────────────────────────────────────────────────

def _upload_supabase(filepath: Path):
    """Upload file ke Supabase Storage. Dipanggil setelah _simpan."""
    if not SUPABASE_KEY or SUPABASE_KEY.startswith("GANTI"):
        log.warning("Supabase key belum diset, skip upload.")
        return

    url     = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{filepath.name}"
    headers = {
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/octet-stream",
        "x-upsert":      "true",   # overwrite kalau sudah ada
    }
    try:
        with open(filepath, "rb") as f:
            resp = requests.put(url, headers=headers, data=f, timeout=60)
        if resp.status_code in (200, 201):
            log.info("Upload Supabase OK: %s", filepath.name)
        else:
            log.error("Upload Supabase gagal %s: %s", resp.status_code, resp.text[:200])
    except Exception as e:
        log.error("Upload Supabase error: %s", e)


def _upload_meta_supabase(meta: dict):
    """Upload last_update.json ke Supabase Storage."""
    if not SUPABASE_KEY or SUPABASE_KEY.startswith("GANTI"):
        return
    url     = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/last_update.json"
    headers = {
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
        "x-upsert":      "true",
    }
    try:
        resp = requests.put(url, headers=headers,
                            data=json.dumps(meta, ensure_ascii=False).encode(),
                            timeout=30)
        if resp.status_code in (200, 201):
            log.info("Upload meta Supabase OK")
        else:
            log.error("Upload meta gagal %s: %s", resp.status_code, resp.text[:200])
    except Exception as e:
        log.error("Upload meta error: %s", e)


def _buat_bucket_jika_belum_ada():
    """Buat bucket Supabase Storage kalau belum ada."""
    if not SUPABASE_KEY or SUPABASE_KEY.startswith("GANTI"):
        return
    url     = f"{SUPABASE_URL}/storage/v1/bucket"
    headers = {"Authorization": f"Bearer {SUPABASE_KEY}", "Content-Type": "application/json"}
    try:
        # Cek dulu apakah bucket sudah ada
        r = requests.get(url, headers=headers, timeout=15)
        buckets = [b.get("name") for b in (r.json() if r.ok else [])]
        if SUPABASE_BUCKET not in buckets:
            payload = {"id": SUPABASE_BUCKET, "name": SUPABASE_BUCKET, "public": False}
            requests.post(url, headers=headers, json=payload, timeout=15)
            log.info("Bucket '%s' dibuat", SUPABASE_BUCKET)
    except Exception as e:
        log.warning("Cek/buat bucket error: %s", e)


def _simpan(df):
    DASHBOARD_DATA.parent.mkdir(parents=True, exist_ok=True)
    log.info("Menyimpan %s ...", DASHBOARD_DATA)

    cols   = [c for c in KOLOM_DISPLAY if c in df.columns]
    df_out = df[cols].copy()

    # Ringkasan Kategori
    kat_df = pd.DataFrame()
    if "Kategori" in df.columns:
        kat_df = (
            df.groupby("Kategori")
            .agg(Jumlah=("No Invoice", "count"),
                 Total_Nominal=("Nominal Invoice", "sum"))
            .reset_index()
            .sort_values("Total_Nominal", ascending=False)
        )

    # Ringkasan Principal
    def _summary_row(x):
        return pd.Series({
            "Total Invoice":        len(x),
            "Lunas":                int(x["_lunas"].sum()),
            "Siap Bayar (count)":   int(x["_siap_bayar"].sum()),
            "Outstanding":          float(x.loc[~x["_lunas"], "Nominal Invoice"].sum()),
            "Nominal Siap Bayar":   float(x.loc[x["_siap_bayar"], "Nominal Invoice"].sum()),
            "Avg Aging (Hari)":     round(x["Aging (Hari)"].mean(), 1) if "Aging (Hari)" in x.columns else 0,
        })

    prin_df = pd.DataFrame()
    if "Principal" in df.columns:
        prin_df = df.groupby("Principal").apply(
            _summary_row, include_groups=False
        ).reset_index()

    # Trend bayar bulanan
    trend_df = pd.DataFrame()
    if "Tgl Payment" in df.columns and "Nominal Bayar" in df.columns:
        paid = df[df["_lunas"] & (df["Nominal Bayar"] > 0)].copy()
        if not paid.empty:
            paid["Period"] = paid["Tgl Payment"].dt.to_period("M")
            trend_df = (
                paid.groupby("Period")["Nominal Bayar"]
                .sum().sort_index().tail(12)
                .reset_index()
            )
            trend_df["Period"] = trend_df["Period"].astype(str)

    with pd.ExcelWriter(DASHBOARD_DATA, engine="openpyxl") as writer:
        df_out.to_excel(writer,  sheet_name="Invoice",             index=False)
        kat_df.to_excel(writer,  sheet_name="Ringkasan Kategori",  index=False)
        prin_df.to_excel(writer, sheet_name="Ringkasan Principal", index=False)
        trend_df.to_excel(writer,sheet_name="Trend Pembayaran",    index=False)
        _style_excel(writer)

    log.info("Tersimpan: %s (%.0f KB)", DASHBOARD_DATA,
             DASHBOARD_DATA.stat().st_size / 1024)

    # Upload ke Supabase
    _buat_bucket_jika_belum_ada()
    _upload_supabase(DASHBOARD_DATA)


def _style_excel(writer):
    from openpyxl.styles import Font, PatternFill, Alignment
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="C0392B")
    ha = Alignment(horizontal="center", vertical="center")
    for ws in writer.sheets.values():
        for cell in ws[1]:
            cell.font, cell.fill, cell.alignment = hf, hfill, ha
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 3, 40)
        ws.freeze_panes = "A2"


def _simpan_meta(n):
    now = datetime.now()
    LAST_UPDATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    meta = {
        "last_update_date": now.strftime("%Y-%m-%d"),
        "last_update_time": now.strftime("%H:%M:%S"),
        "row_count":        n,
        "version":          "1.0",
    }
    LAST_UPDATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(LAST_UPDATE_FILE, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)
    _upload_meta_supabase(meta)


# ──────────────────────────────────────────────────────────────────────────────
#  CONSOLE UI
# ──────────────────────────────────────────────────────────────────────────────

def _banner():
    print("\n" + "=" * 60)
    print("  Update Data – PT Pinus Merah Abadi | FAD")
    print("=" * 60)

def _sukses(n):
    print(f"\n{'='*60}")
    print(f"  ✅  SELESAI  –  {n:,} baris diproses")
    print(f"  Output : {DASHBOARD_DATA}")
    print(f"  Waktu  : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")
    print("\nRefresh browser untuk lihat data terbaru.")
    try: input("\nTekan Enter untuk tutup...")
    except Exception: pass

def _error(msg):
    print(f"\n{'='*60}\n  ❌  GAGAL\n  {msg}\n{'='*60}")
    try: input("\nTekan Enter untuk tutup...")
    except Exception: pass


if __name__ == "__main__":
    main()
